"""Microsoft Excel connector — sync spreadsheet data as markdown tables."""
from __future__ import annotations

import hashlib
import time
from pathlib import Path
from typing import Any, Dict, List

from ..storage.memory import Memory, MemoryStore
from .base import ConfigField, ConnectorPlugin, SyncResult

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _parse_csv(val) -> List[str]:
    if isinstance(val, list):
        return val
    if isinstance(val, str) and val.strip():
        return [t.strip() for t in val.split(",") if t.strip()]
    return []


def _sheet_to_markdown(ws) -> str:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    def _cell(v):
        if v is None:
            return ""
        return str(v).replace("|", "\\|").replace("\n", " ")

    header = rows[0]
    cols = len(header)
    lines = [
        "| " + " | ".join(_cell(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in range(cols)) + " |",
    ]
    for row in rows[1:]:
        padded = list(row) + [None] * (cols - len(row))
        lines.append("| " + " | ".join(_cell(c) for c in padded[:cols]) + " |")

    return "\n".join(lines)


class ExcelConnector(ConnectorPlugin):
    name = "excel"
    display_name = "Microsoft Excel"
    description = "Sync spreadsheet data from Excel files as markdown tables"
    icon = "X"
    category = "Documents"
    setup_guide = "Point to a directory containing .xlsx files. Requires openpyxl (pip install openpyxl)."
    color = "#217346"

    MAX_FILE_SIZE = 50 * 1024 * 1024

    def config_schema(self) -> List[ConfigField]:
        return [
            ConfigField("directory_path", "Directory path", placeholder="/path/to/excel/files", required=True),
            ConfigField("file_pattern", "File pattern", placeholder="*.xlsx", default="*.xlsx"),
            ConfigField("sheet_filter", "Sheet filter", placeholder="e.g. Sheet1, Summary (empty = all)"),
        ]

    @property
    def configured(self) -> bool:
        path = self._config.get("directory_path", "")
        return bool(path and Path(path).is_dir())

    def test_connection(self) -> Dict[str, Any]:
        if not _HAS_OPENPYXL:
            return {"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}

        path = self._config.get("directory_path", "")
        if not path:
            return {"ok": False, "error": "Directory path not set"}
        directory = Path(path)
        if not directory.is_dir():
            return {"ok": False, "error": f"Directory not found: {path}"}

        pattern = self._config.get("file_pattern", "*.xlsx") or "*.xlsx"
        files = list(directory.rglob(pattern))
        if pattern == "*.xlsx":
            files += list(directory.rglob("*.xls"))

        if not files:
            return {"ok": False, "error": f"No Excel files found in {path}"}

        return {"ok": True, "file_count": len(files), "directory": str(directory)}

    def sync(self, store: MemoryStore) -> SyncResult:
        if not self.configured:
            r = SyncResult()
            r.errors.append("Not configured")
            return r

        if not _HAS_OPENPYXL:
            r = SyncResult()
            r.errors.append("openpyxl not installed. Run: pip install openpyxl")
            return r

        result = SyncResult()
        directory = Path(self._config["directory_path"])
        prefix = f"{self.name}/"
        pattern = self._config.get("file_pattern", "*.xlsx") or "*.xlsx"
        sheet_filter = _parse_csv(self._config.get("sheet_filter", ""))

        excel_files = list(directory.rglob(pattern))
        if pattern == "*.xlsx":
            excel_files += list(directory.rglob("*.xls"))

        synced_keys = set()

        for f in excel_files:
            if not f.is_file():
                continue
            if f.stat().st_size > self.MAX_FILE_SIZE:
                result.skipped += 1
                continue
            if f.name.startswith("~"):
                continue

            filename = f.stem

            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            except Exception as e:
                result.errors.append(f"{f.name}: {e}")
                continue

            try:
                for sheet_name in wb.sheetnames:
                    if sheet_filter and sheet_name.lower() not in [s.lower() for s in sheet_filter]:
                        continue

                    ws = wb[sheet_name]
                    md_table = _sheet_to_markdown(ws)
                    if not md_table:
                        result.skipped += 1
                        continue

                    result.total_remote += 1
                    key = f"{prefix}{filename}/{sheet_name}"
                    synced_keys.add(key)

                    content = f"# {filename} - {sheet_name}\n\n{md_table}"
                    content_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
                    mem_tags = [self.name, filename.lower(), sheet_name.lower()]

                    try:
                        existing = store.get(key)
                        if existing.metadata.get("content_hash") == content_hash:
                            result.skipped += 1
                            continue
                        existing.value = content
                        existing.tags = mem_tags
                        existing.metadata["content_hash"] = content_hash
                        existing.metadata["modified"] = f.stat().st_mtime
                        existing.updated_at = time.time()
                        store.set(existing)
                        result.updated += 1
                    except KeyError:
                        mem = Memory(
                            key=key, value=content, tags=mem_tags,
                            metadata={
                                "source": self.name,
                                "content_hash": content_hash,
                                "file_path": str(f),
                                "filename": f.name,
                                "sheet_name": sheet_name,
                                "modified": f.stat().st_mtime,
                            },
                        )
                        store.set(mem)
                        result.added += 1
            finally:
                wb.close()

        for m in store.list():
            if m.key.startswith(prefix) and m.key not in synced_keys:
                store.delete(m.key)
                result.removed += 1

        self._update_sync_stats(len(synced_keys))
        return result
